
import csv
import os
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openai import OpenAI
from openpyxl import load_workbook

# =========================
# 你主要改这里
# =========================
MODEL_NAME = "doubao-seed-2-0-mini-260215"
REPEATS_PER_PROMPT = 3
SLEEP_BETWEEN_CALLS = 1.0
MAX_RETRIES = 3
RETRY_BACKOFF_SECONDS = 2.0

RUN_LABEL = "prompts15_revised"
OUTPUT_ROOT = Path("experiment_outputs") / "doubao" / "prompts15_revised"
SAVE_TXT_BACKUP = True

BASE_URL = "https://ark.cn-beijing.volces.com/api/v3"
PROMPT_SHEET_NAME = "Prompts_15_Revised"

# 默认会在“脚本所在目录”和“当前工作目录”里找这个 Excel。
PROMPT_XLSX_CANDIDATES = [
    Path("self_contained_prompts_15_revised.xlsx"),
    Path(__file__).resolve().parent / "self_contained_prompts_15_revised.xlsx",
    Path("/mnt/data/self_contained_prompts_15_revised.xlsx"),  # 方便当前对话里直接测试
]


@dataclass
class PromptItem:
    base_id: str
    variant: str
    category: str
    note: str
    prompt_id: str
    text: str


VARIANT_SPECS: Sequence[Tuple[str, str, str]] = (
    ("original", "A组：原始自足型Prompt", "original"),
    ("free", "B组：30秒自由压缩（留空）", "free_compressed"),
    ("classical", "C组：文言文版本（留空）", "classical_compressed"),
)


def find_prompt_workbook() -> Path:
    for path in PROMPT_XLSX_CANDIDATES:
        if path.exists():
            return path
    print("错误：未找到 prompt Excel。请把 self_contained_prompts_15_revised.xlsx 放到脚本同目录。")
    raise SystemExit(1)


def get_api_key() -> Optional[str]:
    return os.environ.get("ARK_API_KEY")


def build_client() -> OpenAI:
    api_key = get_api_key()
    if not api_key:
        print("错误：未检测到环境变量 ARK_API_KEY。")
        print("请先在 PowerShell 中执行以下任一方式：")
        print('  当前窗口临时生效：  $env:ARK_API_KEY = "你的_API_Key"')
        print('  永久写入用户环境变量：')
        print('  [Environment]::SetEnvironmentVariable("ARK_API_KEY", "你的_API_Key", "User")')
        raise SystemExit(1)
    return OpenAI(api_key=api_key, base_url=BASE_URL)


def sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(name).strip())
    return cleaned[:120] or "untitled"


def safe_getattr(obj: Any, name: str, default: Any = None) -> Any:
    return getattr(obj, name, default) if obj is not None else default


def usage_to_dict(usage: Any) -> Dict[str, Any]:
    if usage is None:
        return {
            "response_input_token_count": None,
            "response_output_token_count": None,
            "response_total_token_count": None,
            "response_cached_input_token_count": None,
        }

    input_details = safe_getattr(usage, "input_tokens_details", None)
    cached_tokens = safe_getattr(input_details, "cached_tokens", None)
    if cached_tokens is None and isinstance(input_details, dict):
        cached_tokens = input_details.get("cached_tokens")

    return {
        "response_input_token_count": safe_getattr(usage, "input_tokens", None),
        "response_output_token_count": safe_getattr(usage, "output_tokens", None),
        "response_total_token_count": safe_getattr(usage, "total_tokens", None),
        "response_cached_input_token_count": cached_tokens,
    }


def extract_text(response: Any) -> str:
    output_text = safe_getattr(response, "output_text", None)
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()

    texts: List[str] = []
    output_items = safe_getattr(response, "output", []) or []
    for item in output_items:
        content_list = safe_getattr(item, "content", []) or []
        for content in content_list:
            text_val = safe_getattr(content, "text", None)
            if isinstance(text_val, str) and text_val.strip():
                texts.append(text_val.strip())
    return "\n".join(texts).strip()


def load_prompts_from_excel(workbook_path: Path, sheet_name: str) -> List[PromptItem]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"错误：Excel 中未找到工作表 {sheet_name!r}。可用工作表：{wb.sheetnames}")
        raise SystemExit(1)

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("错误：prompt 表为空。")
        raise SystemExit(1)

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    header_index = {name: idx for idx, name in enumerate(header)}

    required = ["ID", "Category", "A组：原始自足型Prompt", "B组：30秒自由压缩（留空）", "C组：文言文版本（留空）", "备注"]
    missing = [name for name in required if name not in header_index]
    if missing:
        print(f"错误：Excel 缺少这些列：{missing}")
        raise SystemExit(1)

    prompts: List[PromptItem] = []
    for row in rows[1:]:
        base_id = str(row[header_index["ID"]]).strip() if row[header_index["ID"]] is not None else ""
        if not base_id:
            continue

        category = str(row[header_index["Category"]] or "").strip()
        note = str(row[header_index["备注"]] or "").strip()

        for variant, col_name, prompt_suffix in VARIANT_SPECS:
            text = str(row[header_index[col_name]] or "").strip()
            if not text:
                continue
            prompts.append(
                PromptItem(
                    base_id=base_id,
                    variant=variant,
                    category=category,
                    note=note,
                    prompt_id=f"{base_id}_{prompt_suffix}",
                    text=text,
                )
            )

    if not prompts:
        print("错误：没有读取到任何 prompt。")
        raise SystemExit(1)
    return prompts


def run_single_request(
    client: OpenAI,
    model_name: str,
    prompt: PromptItem,
    run_index: int,
) -> Dict[str, Any]:
    last_error: Optional[str] = None

    for attempt in range(1, MAX_RETRIES + 1):
        started_at = datetime.now().isoformat(timespec="seconds")
        try:
            response = client.responses.create(
                model=model_name,
                input=prompt.text,
            )

            usage = usage_to_dict(safe_getattr(response, "usage", None))
            output_text = extract_text(response)

            return {
                "base_id": prompt.base_id,
                "variant": prompt.variant,
                "category": prompt.category,
                "note": prompt.note,
                "prompt_id": prompt.prompt_id,
                "run_index": run_index,
                "attempt": attempt,
                "status": "ok",
                "started_at": started_at,
                "model_name": model_name,
                "prompt_text": prompt.text,
                **usage,
                "output_text": output_text,
                "error": "",
                "response_id": safe_getattr(response, "id", None),
            }
        except Exception as exc:
            last_error = repr(exc)
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_BACKOFF_SECONDS * attempt)
            else:
                return {
                    "base_id": prompt.base_id,
                    "variant": prompt.variant,
                    "category": prompt.category,
                    "note": prompt.note,
                    "prompt_id": prompt.prompt_id,
                    "run_index": run_index,
                    "attempt": attempt,
                    "status": "error",
                    "started_at": started_at,
                    "model_name": model_name,
                    "prompt_text": prompt.text,
                    "response_input_token_count": None,
                    "response_output_token_count": None,
                    "response_total_token_count": None,
                    "response_cached_input_token_count": None,
                    "output_text": "",
                    "error": last_error,
                    "response_id": None,
                }

    raise RuntimeError("不应运行到这里。")


def mean_of(key: str, items: List[Dict[str, Any]]) -> Optional[float]:
    vals = [item[key] for item in items if isinstance(item.get(key), (int, float))]
    if not vals:
        return None
    return round(sum(vals) / len(vals), 4)


def build_summary_rows(rows: List[Dict[str, Any]], group_keys: Sequence[str]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = {}
    for row in rows:
        if row.get("status") != "ok":
            continue
        key = tuple(row.get(k) for k in group_keys)
        grouped.setdefault(key, []).append(row)

    summary: List[Dict[str, Any]] = []
    for key_tuple, items in grouped.items():
        record = {k: v for k, v in zip(group_keys, key_tuple)}
        record.update(
            {
                "successful_runs": len(items),
                "mean_response_input_token_count": mean_of("response_input_token_count", items),
                "mean_response_output_token_count": mean_of("response_output_token_count", items),
                "mean_response_total_token_count": mean_of("response_total_token_count", items),
                "mean_response_cached_input_token_count": mean_of("response_cached_input_token_count", items),
            }
        )
        summary.append(record)

    summary.sort(key=lambda x: tuple(str(x.get(k, "")) for k in group_keys))
    return summary


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def maybe_write_excel(run_rows: List[Dict[str, Any]], path: Path) -> str:
    try:
        import pandas as pd
    except Exception:
        return "未生成 xlsx（当前环境未安装 pandas）。CSV 已生成。"

    run_df = pd.DataFrame(run_rows)
    summary_variant = pd.DataFrame(build_summary_rows(run_rows, ["variant"]))
    summary_prompt_family = pd.DataFrame(build_summary_rows(run_rows, ["base_id", "category"]))
    summary_prompt_variant = pd.DataFrame(build_summary_rows(run_rows, ["base_id", "variant", "category", "prompt_id"]))

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        run_df.to_excel(writer, sheet_name="run_level_results", index=False)
        summary_variant.to_excel(writer, sheet_name="summary_by_variant", index=False)
        summary_prompt_family.to_excel(writer, sheet_name="summary_by_prompt", index=False)
        summary_prompt_variant.to_excel(writer, sheet_name="summary_by_prompt_variant", index=False)

        for variant in ["original", "free", "classical"]:
            run_df[run_df["variant"] == variant].to_excel(writer, sheet_name=f"runs_{variant}", index=False)

    return "xlsx 已生成。"


def write_txt_backup(txt_dir: Path, row: Dict[str, Any]) -> Path:
    file_name = f"{sanitize_filename(row['base_id'])}_{sanitize_filename(row['variant'])}_run{int(row['run_index']):02d}.txt"
    txt_path = txt_dir / file_name

    lines = [
        f"base_id: {row.get('base_id', '')}",
        f"variant: {row.get('variant', '')}",
        f"category: {row.get('category', '')}",
        f"prompt_id: {row.get('prompt_id', '')}",
        f"run_index: {row.get('run_index', '')}",
        f"attempt: {row.get('attempt', '')}",
        f"status: {row.get('status', '')}",
        f"started_at: {row.get('started_at', '')}",
        f"model_name: {row.get('model_name', '')}",
        f"response_input_token_count: {row.get('response_input_token_count', '')}",
        f"response_output_token_count: {row.get('response_output_token_count', '')}",
        f"response_total_token_count: {row.get('response_total_token_count', '')}",
        f"response_cached_input_token_count: {row.get('response_cached_input_token_count', '')}",
        f"response_id: {row.get('response_id', '')}",
        f"note: {row.get('note', '')}",
        f"error: {row.get('error', '')}",
        "",
        "=== PROMPT ===",
        str(row.get("prompt_text", "")),
        "",
        "=== OUTPUT ===",
        str(row.get("output_text", "")),
        "",
    ]
    txt_path.write_text("\n".join(lines), encoding="utf-8")
    return txt_path


def write_grouped_outputs(run_dir: Path, rows: List[Dict[str, Any]]) -> None:
    by_variant_dir = run_dir / "by_variant"
    by_variant_dir.mkdir(parents=True, exist_ok=True)

    for variant in ["original", "free", "classical"]:
        variant_rows = [r for r in rows if r.get("variant") == variant]
        if not variant_rows:
            continue

        variant_dir = by_variant_dir / variant
        variant_dir.mkdir(parents=True, exist_ok=True)

        write_csv(variant_dir / f"run_level_results_{variant}.csv", variant_rows)
        write_csv(
            variant_dir / f"summary_{variant}.csv",
            build_summary_rows(variant_rows, ["base_id", "category", "prompt_id"]),
        )


def main() -> int:
    workbook_path = find_prompt_workbook()
    prompts = load_prompts_from_excel(workbook_path, PROMPT_SHEET_NAME)

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    model_tag = sanitize_filename(MODEL_NAME)
    run_dir = OUTPUT_ROOT / f"{RUN_LABEL}_{model_tag}_{timestamp}"
    run_dir.mkdir(parents=True, exist_ok=True)

    run_csv_path = run_dir / "run_level_results_all.csv"
    summary_variant_csv_path = run_dir / "summary_by_variant.csv"
    summary_prompt_csv_path = run_dir / "summary_by_prompt.csv"
    summary_prompt_variant_csv_path = run_dir / "summary_by_prompt_variant.csv"
    run_xlsx_path = run_dir / "results_grouped.xlsx"

    txt_root = run_dir / "txt_backups"
    if SAVE_TXT_BACKUP:
        for variant in ["original", "free", "classical"]:
            (txt_root / variant).mkdir(parents=True, exist_ok=True)

    client = build_client()
    all_rows: List[Dict[str, Any]] = []

    total_jobs = len(prompts) * REPEATS_PER_PROMPT
    finished = 0
    print(f"Excel 来源：{workbook_path.resolve()}")
    print(f"开始运行：{len(prompts)} 个 prompt × {REPEATS_PER_PROMPT} 次 = {total_jobs} 次请求")
    print(f"模型：{MODEL_NAME}")
    print("提示：本脚本会自动把 original / free / classical 分目录保存。")
    print()

    try:
        for prompt in prompts:
            for run_index in range(1, REPEATS_PER_PROMPT + 1):
                finished += 1
                print(f"[{finished}/{total_jobs}] {prompt.prompt_id} 第 {run_index} 次运行中...")
                row = run_single_request(
                    client=client,
                    model_name=MODEL_NAME,
                    prompt=prompt,
                    run_index=run_index,
                )
                all_rows.append(row)

                write_csv(run_csv_path, all_rows)
                write_csv(summary_variant_csv_path, build_summary_rows(all_rows, ["variant"]))
                write_csv(summary_prompt_csv_path, build_summary_rows(all_rows, ["base_id", "category"]))
                write_csv(
                    summary_prompt_variant_csv_path,
                    build_summary_rows(all_rows, ["base_id", "variant", "category", "prompt_id"]),
                )
                write_grouped_outputs(run_dir, all_rows)

                if SAVE_TXT_BACKUP:
                    write_txt_backup(txt_root / str(row["variant"]), row)

                if SLEEP_BETWEEN_CALLS > 0 and finished < total_jobs:
                    time.sleep(SLEEP_BETWEEN_CALLS)
        excel_message = maybe_write_excel(all_rows, run_xlsx_path)
    finally:
        try:
            client.close()
        except Exception:
            pass

    print("\n=== 完成 ===")
    print(f"结果目录：{run_dir.resolve()}")
    print(f"总表 CSV：{run_csv_path.resolve()}")
    print(f"按条件汇总 CSV：{summary_variant_csv_path.resolve()}")
    print(f"按题目汇总 CSV：{summary_prompt_csv_path.resolve()}")
    print(f"按题目+条件汇总 CSV：{summary_prompt_variant_csv_path.resolve()}")
    print(f"Excel：{run_xlsx_path.resolve()}  （{excel_message}）")
    if SAVE_TXT_BACKUP:
        print(f"TXT 备份根目录：{txt_root.resolve()}")

    print("\n你最关心的点：")
    print("1) 会自动从 self_contained_prompts_15_revised.xlsx 的 Prompts_15_Revised 读取 15 组 prompt。")
    print("2) 每组会展开为 original / free / classical 三种条件，总共 45 个 prompt。")
    print("3) 结果会同时保存总表和分条件目录，不用你再手动整理。")
    print("4) 每次请求后都会立刻落盘 CSV 和 txt，半路中断时前面已完成的结果仍会保留。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
